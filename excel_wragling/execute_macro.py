import win32com.client
import os
import openpyxl as px
import os
from typing import List
import time
from retrying import retry
CURRENT_DIR = os.getcwd()
FNAME="MUST_CHECK.txt"


@retry(stop_max_attempt_number=3)
def execute_macro(fname: str, macro_name: str):
    # open file and execute a excel-macro
    print("{}を実行".format(macro_name))
    excel = win32com.client.Dispatch("Excel.Application")  # インスタンス生成
    excel.Visible = False  # エクセルを表示する設定（Falseにすれば非表示で実行される）
    excel.Workbooks.Open(Filename=os.path.join(
        CURRENT_DIR, fname), ReadOnly=False)  # ブックを読み取り専用で開く
    excel.Application.Run(macro_name)  # マクロ名を指定して実行（引数なしの場合マクロ名のみで実行可能）
    # ブックを保存して閉じる（SaveChangesを0にすると保存せず閉じる）
    excel.Workbooks(1).Close(SaveChanges=1)
    excel.Application.Quit()  # 終了


def remove_needless_cell(fname: str, sheet_name="CVSSレベル取得"):
	# delete neddles line from excel
	print("不要な行を削除")
	wb = px.load_workbook(fname, data_only=True, keep_vba=True)
    # titles = wb.get_sheet_names()
	sheet = wb.get_sheet_by_name(sheet_name)
	row_: int = 0
	col_: int = 12
	index_list = []
	max=sheet.max_row
	print(max)
	#time.sleep(2)
	while  row_<=max:
		row_ += 1
		content=sheet.cell(row=row_, column=col_).value
		if content:
			#print(content)
			if not content.lower().__contains__("cvss"):
				with open(FNAME,"a",encoding="utf-8")as writer:
					print("要確認:{} {}行目!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!".format(FNAME,row_))
					print(content)
					print("")
					writer.write(fname+"\t"+str(row_)+"行目\t"+content+"\n")

			row_ += 2
			index_list.append(row_)

	#print(index_list)
	wb.close()
	# data_onlyにして書き換え
	wb = px.load_workbook(fname, keep_vba=True)
	sheet = wb.get_sheet_by_name(sheet_name)
	for row_ in index_list:
		sheet.cell(row=row_, column=col_).value = None
	wb.save(os.path.join(CURRENT_DIR, fname))


if __name__ == "__main__":
    # macro実行
    #current-dirのxlmxをすべて実行する
	start=time.time()
	fname_list = os.listdir()
	for fname in fname_list:
		if fname.endswith(".xlsm"):
   			print("#######################################s")
   			print(fname)
   			print("#######################################s")
   			# execute_macro(fname, "データ削除")
   			execute_macro(fname, "バッチログコピー")
   			remove_needless_cell(fname)
   			execute_macro(fname,"ボタン2_Click")
	print("実行時間")
	print(time.time()-start)
